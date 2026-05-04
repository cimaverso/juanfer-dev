import csv
from openpyxl import load_workbook


def parse_file(file):
    """
    Convierte archivo CSV o Excel a lista de diccionarios.
    """

    filename = file.filename.lower()

    if filename.endswith(".csv"):
        content = file.file.read().decode("utf-8").splitlines()
        reader = csv.DictReader(content)
        return list(reader)

    elif filename.endswith(".xlsx"):
        wb = load_workbook(file.file)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

        return rows

    else:
        raise ValueError("Formato no soportado")